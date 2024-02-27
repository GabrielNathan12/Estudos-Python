from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER/ 'workbook.xlsx'

workbook = Workbook()
workbook.create_sheet('Primeira página',0)
worksheet: Worksheet = workbook.active

# Remover uma planilha
#workbook.remove(workbook['Sheet'])

worksheet.cell(1,1, 'Nome')
worksheet.cell(1,2, 'Idade')
worksheet.cell(1,3, 'Nota')

students = [
    #Nome      #Idade  #Nota
    ['João',   15,     6.8],
    ['Maria',  12,     9.6],
    ['Luiz',   16,     8.9],
    ['Gabriel',16,     10]
]

'''
for i, student_row in enumerate(students, start=2):
    for j, student_col in enumerate(student_row,start=1):
        worksheet.cell(i, j, student_col)

'''


for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)