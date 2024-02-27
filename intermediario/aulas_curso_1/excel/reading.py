from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER/ 'workbook.xlsx'

workbook = load_workbook(WORKBOOK_PATH)

worksheet: Worksheet = workbook['Primeira página']

row: tuple[Cell]

for row in worksheet.iter_rows(min_row=2):
    for col in row:
        print(col.value, end='\t')

        if col.value == 'Maria':
            worksheet.cell(col.row, 2, 17)
    print()

#worksheet['B3'].value = 16

workbook.save(WORKBOOK_PATH)